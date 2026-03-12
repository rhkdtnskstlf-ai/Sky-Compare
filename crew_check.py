import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

# 페이지 설정
st.set_page_config(layout="wide", page_title="Crew 분석 시스템 Pro")

# --- 커스텀 CSS (너비 확보 및 가독성 개선) ---
st.markdown("""
    <style>
    /* 테이블 내 텍스트가 세로로 쪼개지지 않도록 설정 */
    .stTable td, .stTable th {
        white-space: nowrap !important;
        font-size: 0.85rem !important;
        padding: 5px !important;
    }
    .group-card { background-color: #ffffff; border: 1px solid #e9ecef; border-left: 6px solid #fd7e14; padding: 20px; margin-bottom: 20px; border-radius: 12px; box-shadow: 2px 4px 12px rgba(0,0,0,0.08); }
    .move-group-card { background-color: #f0f4ff; border: 1px solid #dbe4ff; border-left: 6px solid #4c6ef5; padding: 20px; margin-bottom: 20px; border-radius: 12px; }
    .move-title { font-weight: 800; color: #1c7ed6; margin-bottom: 10px; font-size: 1.1em; }
    .flight-title { color: #1e293b; font-size: 1.3em; font-weight: 800; margin-bottom: 15px; border-bottom: 2px solid #f1f3f5; padding-bottom: 10px; }
    .item-container { display: flex; align-items: center; padding: 10px 14px; margin: 8px 0; border-radius: 8px; font-size: 0.95rem; line-height: 1.5; font-weight: 500; }
    .bg-swap { background-color: #f8f9fa; border-left: 4px solid #74c0fc; }
    .bg-in { background-color: #f3fcf3; border-left: 4px solid #40c057; }
    .bg-out { background-color: #fff5f5; border-left: 4px solid #ff8787; }
    .bg-info { background-color: #fff9db; border-left: 4px solid #fab005; }
    .badge { padding: 2px 8px; border-radius: 4px; font-size: 0.85em; font-weight: 800; margin-right: 12px; color: white; min-width: 45px; text-align: center; }
    .badge-swap { background-color: #1c7ed6; }
    .badge-in { background-color: #2f9e44; }
    .badge-out { background-color: #e03131; } /* CXL 배지 색상 유지 */
    .badge-info { background-color: #f08c00; }
    .flight-header { background-color: #f1f3f5; padding: 10px 15px; border-radius: 8px; font-weight: bold; margin-top: 15px; border-left: 4px solid #495057; }
    </style>
    """, unsafe_allow_html=True)

# --- 유틸리티 함수 ---
def normalize_id(x):
    if pd.isna(x) or str(x).strip() == "": return ""
    text = str(x).strip()
    if text.endswith('.0'): text = text[:-2]
    return text.upper()

def format_time_display(val):
    if pd.isna(val) or val == "": return "-"
    if hasattr(val, 'strftime'): return val.strftime("%H:%M")
    try:
        str_val = str(val).strip()
        if len(str_val) >= 10: return pd.to_datetime(str_val).strftime("%H:%M")
        return str_val[:5]
    except: return str(val)

def is_layover_left(cell):
    try:
        if not cell.fill or cell.fill.fill_type is None: return False
        color_obj = cell.fill.start_color
        rgb_val = str(color_obj.rgb).upper() if color_obj.rgb else ""
        if rgb_val in ["", "00000000", "FFFFFFFF", "NONE"]: return False
        if "FFFF00" in rgb_val or "FFFFCC" in rgb_val or "FFFFE0" in rgb_val: return False
        return color_obj.index != 64
    except: return False

# --- 데이터 로더 ---
def load_crew_left(file, sheet_name):
    wb = openpyxl.load_workbook(file, data_only=False) 
    ws = wb[sheet_name]
    data = []
    for r in range(8, 500):
        c_id_val = ws[f"A{r}"].value
        if c_id_val is None: continue
        name_cell = ws[f"B{r}"] 
        data.append({
            "CrewID": normalize_id(c_id_val),
            "CrewName": str(name_cell.value).strip() if name_cell.value else "Unknown",
            "Arr Flt": str(ws[f"G{r}"].value).strip().upper() if ws[f"G{r}"].value else "OPEN",
            "Arr Time": ws[f"H{r}"].value,
            "Dep Flt": str(ws[f"J{r}"].value).strip().upper() if ws[f"J{r}"].value else "OPEN", 
            "Dep Time": ws[f"K{r}"].value, 
            "is_layover": is_layover_left(name_cell)
        })
    df = pd.DataFrame(data)
    df = df[df["CrewID"] != ""]
    return df.sort_values(by=["Arr Time", "Arr Flt", "CrewName"]).reset_index(drop=True)

def load_crew_right(file):
    try:
        df = pd.read_excel(file, header=2, usecols="D:E,O:P,Q:R", engine='openpyxl')
        df.columns = ["CrewID", "CrewName", "Arr Flt", "Arr Time", "Dep Flt", "Dep Time"]
        df["CrewID"] = df["CrewID"].apply(normalize_id)
        df["Date_Only"] = pd.to_datetime(df["Arr Time"], dayfirst=True, errors='coerce').dt.date
        df = df.dropna(subset=["CrewID"])
        return df.sort_values(by=["Arr Time", "Arr Flt", "CrewName"]).reset_index(drop=True)
    except: return None

# --- 사이드바 ---
with st.sidebar:
    st.header("⚙️ 분석 설정")
    show_layover_only = st.checkbox("🏨 연박 인원만 보기", value=False)
    if st.button("🔄 데이터 초기화"):
        st.cache_data.clear()
        st.rerun()

st.title("✈️ Crew 명단 통합 분석 Pro")
up_l, up_r = st.columns(2)
df_l = df_r = None

with up_l:
    f_l = st.file_uploader("기존 명단 (Old)", type=["xlsx"])
    if f_l:
        wb_l = openpyxl.load_workbook(f_l, read_only=True)
        sh_l = st.selectbox("시트 선택", wb_l.sheetnames)
        df_l = load_crew_left(f_l, sh_l)

with up_r:
    f_r = st.file_uploader("신규 명단 (New)", type=["xlsx"])
    if f_r:
        df_r_raw = load_crew_right(f_r)
        if df_r_raw is not None:
            u_dates = sorted([d for d in df_r_raw["Date_Only"].unique() if pd.notna(d)])
            sel_d = st.selectbox("도착일 선택", u_dates)
            next_d = sel_d + timedelta(days=1)
            today_data = df_r_raw[df_r_raw["Date_Only"] == sel_d].copy()
            next_day_ids = set(df_r_raw[df_r_raw["Date_Only"] == next_d]["CrewID"])
            today_data["is_layover"] = today_data["CrewID"].apply(lambda x: x in next_day_ids)
            df_r = today_data.sort_values(by=["Arr Time", "Arr Flt", "CrewName"]).reset_index(drop=True)

# --- 분석 섹션 ---
if df_l is not None and df_r is not None:
    if show_layover_only:
        df_l = df_l[df_l['is_layover'] == True].copy()
        df_r = df_r[df_r['is_layover'] == True].copy()

    st.divider()
    view_l, view_center, view_r = st.columns([1.2, 1.6, 1.2])

    def display_list(container, df, title):
        with container:
            st.subheader(title)
            if df.empty:
                st.write("데이터 없음")
                return
            for flt in df["Arr Flt"].unique():
                gp = df[df["Arr Flt"] == flt]
                st.markdown(f"<div class='flight-header'>{flt} ({len(gp)}명)</div>", unsafe_allow_html=True)
                disp = gp.copy()
                disp["🏨"] = disp["is_layover"].map({True: "✅", False: ""})
                disp["이름(ID)"] = disp["CrewName"] + "(" + disp["CrewID"] + ")"
                disp["도착"] = disp["Arr Time"].apply(format_time_display)
                disp["출발"] = disp["Dep Flt"] + " (" + disp["Dep Time"].apply(format_time_display) + ")"
                st.table(disp[["🏨", "도착", "이름(ID)", "출발"]])

    display_list(view_l, df_l, "⬅️ 기존 명단")
    display_list(view_r, df_r, "➡️ 신규 명단")

    with view_center:
        st.markdown("<h2 style='text-align: center; margin-bottom: 25px;'>📋 통합 변경 리포트</h2>", unsafe_allow_html=True)
        
        all_merged = pd.merge(df_l, df_r, on="CrewID", how="outer", suffixes=('_old', '_new'))
        moved_crew = all_merged[all_merged['Arr Flt_old'].notna() & all_merged['Arr Flt_new'].notna() & (all_merged['Arr Flt_old'] != all_merged['Arr Flt_new'])].copy()
        processed_ids = set(moved_crew['CrewID'].tolist())
        
        if not moved_crew.empty:
            move_groups = moved_crew.groupby(['Arr Flt_old', 'Arr Flt_new'])
            for (old_f, new_f), group in move_groups:
                names_html = " ".join([f"<span class='badge' style='background-color:#4c6ef5; display:inline-block; margin-bottom:5px;'>{n}</span>" for n in group['CrewName_new']])
                st.markdown(f"<div class='move-group-card'><div class='move-title'>🚚 편수 이동: {old_f} ➔ {new_f}</div>{names_html}</div>", unsafe_allow_html=True)

        sorted_flts = df_l.sort_values(by=["Arr Time", "Arr Flt"])["Arr Flt"].unique().tolist()
        new_only_flts = [f for f in df_r["Arr Flt"].unique() if f not in sorted_flts]
        all_flts_ordered = sorted_flts + new_only_flts
        
        for flt in all_flts_ordered:
            curr_old = df_l[(df_l["Arr Flt"] == flt) & (~df_l["CrewID"].isin(processed_ids))]
            curr_new = df_r[(df_r["Arr Flt"] == flt) & (~df_r["CrewID"].isin(processed_ids))]
            
            old_ids = set(curr_old["CrewID"])
            new_ids = set(curr_new["CrewID"])
            out_ids = old_ids - new_ids
            in_ids = new_ids - old_ids
            stay_ids = old_ids & new_ids
            
            items_html = []
            rem_list = curr_old[curr_old["CrewID"].isin(out_ids)].to_dict('records')
            add_list = curr_new[curr_new["CrewID"].isin(in_ids)].to_dict('records')
            
            match_cnt = min(len(rem_list), len(add_list))
            for _ in range(match_cnt):
                r, a = rem_list.pop(0), add_list.pop(0)
                items_html.append(f"<div class='item-container bg-swap'><span class='badge badge-swap'>교체</span> {r['CrewName']} ➔ <b>{a['CrewName']}</b></div>")

            # --- OUT에서 CXL로 변경된 부분 ---
            for r in rem_list:
                items_html.append(f"<div class='item-container bg-out'><span class='badge badge-out'>CXL</span> {r['CrewName']}</div>")
            
            for a in add_list:
                items_html.append(f"<div class='item-container bg-in'><span class='badge badge-in'>IN</span> {a['CrewName']}</div>")

            for sid in stay_ids:
                o_r = curr_old[curr_old["CrewID"] == sid].iloc[0]
                n_r = curr_new[curr_new["CrewID"] == sid].iloc[0]
                sub = []
                
                if o_r['is_layover'] != n_r['is_layover']:
                    sub.append("연박 변경" if n_r['is_layover'] else "연박 해제")
                
                t_old, t_new = format_time_display(o_r['Arr Time']), format_time_display(n_r['Arr Time'])
                if t_old != t_new:
                    sub.append(f"도착시간: {t_old} ➔ {t_new}")
                
                d_f_old, d_f_new = str(o_r.get('Dep Flt', '-')), str(n_r.get('Dep Flt', '-'))
                d_t_old, d_t_new = format_time_display(o_r['Dep Time']), format_time_display(n_r['Dep Time'])
                
                if d_f_old != d_f_new or d_t_old != d_t_new:
                    sub.append(f"출발: {d_f_old} {d_t_old} ➔ <b>{d_f_new} {d_t_new}</b>")

                if sub:
                    items_html.append(f"<div class='item-container bg-info'><span class='badge badge-info'>변경</span> <b>{n_r['CrewName']}</b>: {' / '.join(sub)}</div>")

            if items_html:
                st.markdown(f"<div class='group-card'><div class='flight-title'>✈️ {flt}</div>{''.join(items_html)}</div>", unsafe_allow_html=True)

    st.success("✅ 업데이트 완료: 이름 제외 인원에 대해 'CXL' 배지가 적용되었습니다.")
else:
    st.info("💡 파일을 업로드하여 분석을 시작하세요.")
